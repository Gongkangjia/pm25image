from PIL import Image, ImageDraw, ImageFont
import numpy as np
import pandas as pd
import arrow
from pathlib import Path
from loguru import logger
from openpyxl import load_workbook
from config import STATIONS_CNEMC


class GeneratorBase:
    def __init__(self, df):
        self.df = df.reset_index()
        self.df_ts = df
        self.root = Path(__file__).absolute().parent
        self.output_dir = self.root.joinpath("output")
        self.time_h = arrow.now().shift(minutes=-30)
        self.rt_count = df.loc[df.DATETIME.dt.hour ==
                               self.time_h.hour, "PM2_5"].count()


class ImageGenerator(GeneratorBase):
    def __init__(self, df, is_jn=False):
        super(ImageGenerator, self).__init__(df)
        self.is_jn = is_jn
        self.df = self.preprocessing(df)
        self.ax = None
        self.columns_width = np.array(
            [8, 12, 10, 10, 10, 10, 10, 10, 10, 10]) * 60
        self.row_height = [180, 360, 180, *[180] * 16]
        self.fonesize = 140
        self.font = ImageFont.truetype(
            font=str(self.root / 'static/kjgong.ttf'), size=self.fonesize)
        self.fontbd = ImageFont.truetype(
            font=str(self.root / 'static/Times New Roman Bold.ttf'), size=self.fonesize)
        self.linewidth = 6
        self.spacing = self.fonesize * 0.1
        self.start = (10, 10)

    def preprocessing(self, df):
        df = df.set_index("DATETIME")
        df = df.rename({"PM2_5": "PM25"}, axis=1)
        realtime = df.loc[self.time_h.format("YYYY-MM-DD HH:00")]
        day = df.groupby("NAME").mean()
        day["MDA8"] = df.groupby("NAME")["O3_8H"].max()
        day.at["南京", "MDA8"] = day.loc[STATIONS_CNEMC.keys(), "MDA8"].mean()
        res_df = realtime.merge(day, left_on="NAME",
                                right_index=True, suffixes=["_RT", "_DAY"])
        res_df = res_df.set_index("NAME")
        return res_df

    def draw_rec_text(self, loc, text, fill="black", font=None, rfill=None):
        if font is None:
            font = self.font
        else:
            font = self.fontbd

        x, y, w, h = self.get_location(*loc)

        self.drawer.rectangle(xy=((x, y), (x + w, y + h)), fill=rfill,
                              outline="black", width=self.linewidth)

        text = self.get_str(text)
        for i, t in enumerate(text):
            fsize = font.getsize(t)
            height_offset = (i - len(text) / 2 + 0.5) * \
                (self.spacing + self.fonesize)
            self.drawer.text(xy=(x + w / 2 - fsize[0] / 2, y + h / 2 - fsize[1] / 2 + height_offset),
                             text=t,
                             fill=fill,
                             font=font
                             )

    def get_location(self, row_start, col_start, row_end=None, col_end=None, ):
        if not (row_end or col_end):
            row_end = row_start
            col_end = col_start
        x = sum(self.columns_width[:col_start]) + self.start[0]
        y = sum(self.row_height[:row_start]) + self.start[1]
        w = sum(self.columns_width[col_start:col_end + 1])
        h = sum(self.row_height[row_start:row_end + 1])
        return x, y, w, h

    @staticmethod
    def get_str(x):
        if isinstance(x, str):
            return [x]
        if isinstance(x, list):
            return x
        try:
            return [str(round(x))]
        except Exception as e:
            return ["-"]

    def run(self, image_time=None):
        if self.rt_count < 10:
            return False
        if image_time is None:
            image_time = arrow.now().shift(minutes=-30)
        if self.is_jn:
            self.output = self.output_dir.joinpath(
                f"JN_{image_time.format('YYYY-MM-DDTHH')}.png")
        else:
            self.output = self.output_dir.joinpath(
                f"{image_time.format('YYYY-MM-DDTHH')}.png")

        logger.info("创建画布")
        width = sum(self.columns_width)
        height = sum(self.row_height)
        self.ax = Image.new(mode="RGB", size=(
            width + 20, height + 20), color="white")
        logo = Image.open(self.root.joinpath("static").joinpath("logo.v1.png"))
        logo = logo.resize((width, int(logo.height*width/logo.width)))
        self.ax.paste(logo)
        self.drawer = ImageDraw.Draw(self.ax)

        logger.info("画大矩形")
        x, y, w, h = *self.start, width, height
        self.drawer.rectangle(xy=((x, y), (x + w, y + h)),
                              fill=None, outline="black", width=12)

        self.draw_rec_text((0, 0, 2, 0), "序号")
        self.draw_rec_text((0, 1, 2, 1), "点位")
        # datetime = arrow.get(self.datetime_tag_path.read_text(), "YYYY-MM-DDTHH")
        self.draw_rec_text((0, 2, 0, 9), image_time.format("YYYY-MM-DD HH:00"))

        self.draw_rec_text((1, 2, 1, 3), ["PM2.5", "（微克/立方米）"])
        self.draw_rec_text((2, 2), "实时")
        self.draw_rec_text((2, 3), "当日累计")

        self.draw_rec_text((1, 4, 1, 5), ["PM10", "（微克/立方米）"])
        self.draw_rec_text((2, 4), "实时")
        self.draw_rec_text((2, 5), "当日累计")

        self.draw_rec_text((1, 6, 1, 7), ["NO2", "（微克/立方米）"])
        self.draw_rec_text((2, 6), "实时")
        self.draw_rec_text((2, 7), "当日累计")

        self.draw_rec_text((1, 8, 1, 9), ["O3", "（微克/立方米）"])
        self.draw_rec_text((2, 8), "实时")
        self.draw_rec_text((2, 9), "MDA8")
        # self.ax = self.ax.resize((2000, int(self.ax.size[1] / self.ax.size[0] * 2000)), Image.ANTIALIAS)
        # self.ax = self.ax.quantize(colors=128, method=2)

        # 开始画数据
        for station_index, station_name, in enumerate(STATIONS_CNEMC.keys()):
            if station_name == "彩虹桥" and self.is_jn:
                self.draw_rec_text((station_index + 3, 0),
                                   station_index + 1, font=True, rfill="#ffff00")
                self.draw_rec_text((station_index + 3, 1),
                                   station_name, rfill="#ffff00")
            else:
                self.draw_rec_text((station_index + 3, 0), station_index + 1)
                self.draw_rec_text((station_index + 3, 1), station_name)
        if self.is_jn:
            self.draw_rec_text((16, 0, 16, 1), "全市", rfill="#ffff00")
        else:
            self.draw_rec_text((16, 0, 16, 1), "全市")

        self.draw_rec_text((17, 0, 17, 1), "无锡")
        self.draw_rec_text((18, 0, 18, 1), "苏州")

        for species_index, species in enumerate(["PM25_RT", "PM25_DAY", "PM10_RT", "PM10_DAY",
                                                 "NO2_RT", "NO2_DAY", "O3_RT", "MDA8"]):
            # 画国控点
            df_species = self.df.loc[STATIONS_CNEMC.keys(), species].to_frame(
                name="VALUE")
            df_species["COLOR"] = "black"
            df_species["FONT"] = None
            df_species.loc["彩虹桥", "FONT"] = True

            df_species["RFILL"] = None
            df_species.loc["彩虹桥", "RFILL"] = "#ffff00"

            df_species.loc[df_species["VALUE"].isin(
                df_species["VALUE"].nlargest(3)), "COLOR"] = "red"
            df_species = df_species.reset_index()
            for index, row_data in df_species.iterrows():
                # print(index, species_index)
                # logger.info(df_species)
                if self.is_jn and row_data.FONT:
                    self.draw_rec_text((index + 3, species_index + 2),
                                       row_data.VALUE,
                                       fill=row_data.COLOR,
                                       font=True,
                                       rfill=row_data.RFILL
                                       )
                else:
                    self.draw_rec_text(
                        (index + 3, species_index + 2), row_data.VALUE, fill=row_data.COLOR)

            # 画城市
            df_species = self.df.loc[["南京", "无锡", "苏州"],
                                     species].to_frame(name="VALUE")
            df_species["COLOR"] = "black"
            df_species["FONT"] = None
            df_species.loc["南京", "FONT"] = True
            df_species["RFILL"] = None
            df_species.loc["南京", "RFILL"] = "#ffff00"

            df_species.loc[df_species["VALUE"] <
                           df_species.at["南京", "VALUE"], "COLOR"] = "green"
            df_species = df_species.reset_index()
            for index, row_data in df_species.iterrows():
                print(index, species_index)
                if self.is_jn and row_data.FONT:
                    self.draw_rec_text((index + 16, species_index + 2),
                                       row_data.VALUE,
                                       fill=row_data.COLOR,
                                       font=True,
                                       rfill=row_data.RFILL
                                       )
                else:
                    self.draw_rec_text(
                        (index + 16, species_index + 2), row_data.VALUE, fill=row_data.COLOR)

        logger.info("绘制完整正在保存=>{}", self.output)
        self.ax = self.ax.resize(
            (2000, int(self.ax.size[1] / self.ax.size[0] * 2000)), Image.ANTIALIAS)
        self.ax.save(self.output)
        return self.output


class ExcelGenerator(GeneratorBase):
    def __init__(self, df):
        super(ExcelGenerator, self).__init__(df)
        self.output = self.output_dir.joinpath(
            f"{self.time_h.format('YYYY-MM-DDTHH')}.xlsx")
        self.wb = load_workbook(self.root.joinpath(
            "static").joinpath("template8.xlsx"))

    @staticmethod
    def _write_row(ws, row_no, data_list):
        # print(data_list)
        for col, i in enumerate(data_list):
            # print(dict(column=col, row=row_no, value=i))
            _ = ws.cell(column=col + 1, row=row_no, value=i)

    def run(self, image_time=None):
        if self.rt_count < 10:
            return False

        if image_time is None:
            image_time = arrow.now().shift(minutes=-30)
        ws = self.wb["DATA"]
        station_list = self.df.NAME.unique()
        for station_index, station_name in enumerate(station_list):
            station_df = self.df.loc[self.df.NAME == station_name]
            station_df = station_df.set_index("DATETIME")
            station_df = station_df.reindex(pd.date_range(
                start=station_df.index.min(), freq="H", periods=24, name="DATETIME"))
            station_df["NAME"] = station_df.NAME.fillna(station_name)
            station_df = station_df.reset_index()
            for index, d in station_df.iterrows():
                data_list = [d["NAME"], d.DATETIME.strftime(
                    "%Y-%m-%dT%H:%M"), d["PM2_5"], d["PM10"], d["NO2"], d["O3"]]
                print(data_list)
                self._write_row(ws, station_index*24 +
                                index + 2, data_list=data_list)

        ws = self.wb["IMAGE"]
        ws.cell(column=3, row=1, value=image_time.format("YYYY-MM-DDTHH:00"))
        self.wb.save(self.output)
        return self.output


class JiangningImage(GeneratorBase):
    def __init__(self, df, is_jn=False):
        super(JiangningImage, self).__init__(df)
        self.is_jn = is_jn
        self.df = self.preprocessing(df)
        self.ax = None
        self.columns_width = np.array([8, 12,
                                       6, 6, 10, 6,
                                       6, 6, 10, 6,
                                       6, 6, 10, 6,
                                       6, 6, 10, 6,

                                       ]) * 60
        self.row_height = np.array([18, 36, 18,
                                    18, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18,
                                    18
                                    ])*12

        self.fonesize = 140
        self.font = ImageFont.truetype(font=str(self.root / 'static/kjgong.ttf'),
                                       size=self.fonesize)
        self.linewidth = 6
        self.spacing = self.fonesize * 0.1
        self.start = (10, 10)

    def preprocessing(self, df):
        df = df.set_index("DATETIME")
        df = df.rename({"PM2_5": "PM25"}, axis=1)
        realtime = df.loc[self.time_h.format("YYYY-MM-DD HH:00")]
        day = df.groupby("NAME").mean()
        day["MDA8"] = df.groupby("NAME")["O3_8H"].max()
        day.at["南京", "MDA8"] = day.loc[STATIONS_CNEMC.keys(), "MDA8"].mean()
        res_df = realtime.merge(day, left_on="NAME",
                                right_index=True, suffixes=["_RT", "_DAY"])
        res_df = res_df.round(0).set_index("NAME")
        rank = res_df.loc[STATIONS_CNEMC.keys()].rank(method="min").rename(
            lambda x: f"{x}_RANK", axis=1)
        res_df = res_df.join(rank)
        return res_df

    def draw_rec_text(self, loc, text, fill="black", rfill=None, **kwagrs):
        x, y, w, h = self.get_location(*loc)

        self.drawer.rectangle(xy=((x, y), (x + w, y + h)), fill=rfill,
                              outline="black", width=self.linewidth)

        text = self.get_str(text)
        for i, t in enumerate(text):
            fsize = self.font.getbbox(t)[-2:]
            height_offset = (i - len(text) / 2 + 0.5) * \
                (self.spacing + self.fonesize)
            self.drawer.text(xy=(x + w / 2 - fsize[0] / 2, y + h / 2 - fsize[1] / 2 + height_offset),
                             text=t,
                             fill=fill,
                             font=self.font,
                             **kwagrs
                             )

    def get_location(self, row_start, col_start, row_end=None, col_end=None, ):
        if not (row_end or col_end):
            row_end = row_start
            col_end = col_start
        x = sum(self.columns_width[:col_start]) + self.start[0]
        y = sum(self.row_height[:row_start]) + self.start[1]
        w = sum(self.columns_width[col_start:col_end + 1])
        h = sum(self.row_height[row_start:row_end + 1])
        return x, y, w, h

    @staticmethod
    def get_str(x):
        if isinstance(x, str):
            return [x]
        if isinstance(x, list):
            return x
        try:
            return [str(round(x))]
        except Exception as e:
            return [""]

    def run(self):
        if self.rt_count < 10:
            return False

        output = self.output_dir.joinpath(
            f"JN_{self.time_h.format('YYYY-MM-DDTHH')}.png")
        logger.info("创建画布")
        width = sum(self.columns_width)
        height = sum(self.row_height)

        self.ax = Image.new(mode="RGB", size=(
            width + 20, height + 20), color="white")

        logo = Image.open(self.root.joinpath("static").joinpath("logo.v1.png"))
        logo = logo.resize((width, int(logo.height*width/logo.width)))
        logger.info(f"输出图片大小=>{width},{height}")
        logger.info(f"logo改变大小为=>{logo.width},{logo.height}")
        self.ax.paste(logo, (
            int(width/2-logo.width/2),
            int(height/2-logo.height/2)-300
        )
        )
        self.drawer = ImageDraw.Draw(self.ax)

        logger.info("画大矩形")
        x, y, w, h = *self.start, width, height
        self.drawer.rectangle(xy=((x, y), (x + w, y + h)),
                              fill=None, outline="black", width=12)

        self.draw_rec_text((0, 0, 2, 0), "序号")
        self.draw_rec_text((0, 1, 2, 1), "点位")
        # datetime = arrow.get(self.datetime_tag_path.read_text(), "YYYY-MM-DDTHH")
        self.draw_rec_text(
            (0, 2, 0, 17), self.time_h.format("YYYY-MM-DD HH:00"))

        self.draw_rec_text((1, 2, 1, 5), ["PM2.5", "（微克/立方米）"])
        self.draw_rec_text((2, 2), "实时")
        self.draw_rec_text((2, 3), "排名")
        self.draw_rec_text((2, 4), "当日累计")
        self.draw_rec_text((2, 5), "排名")

        self.draw_rec_text((1, 6, 1, 9), ["PM10", "（微克/立方米）"])
        self.draw_rec_text((2, 6), "实时")
        self.draw_rec_text((2, 7), "排名")
        self.draw_rec_text((2, 8), "当日累计")
        self.draw_rec_text((2, 9), "排名")

        self.draw_rec_text((1, 10, 1, 13), ["NO2", "（微克/立方米）"])
        self.draw_rec_text((2, 10), "实时")
        self.draw_rec_text((2, 11), "排名")
        self.draw_rec_text((2, 12), "当日累计")
        self.draw_rec_text((2, 13), "排名")

        self.draw_rec_text((1, 14, 1, 17), ["O3", "（微克/立方米）"])
        self.draw_rec_text((2, 14), "实时")
        self.draw_rec_text((2, 15), "排名")
        self.draw_rec_text((2, 16), "当日累计")
        self.draw_rec_text((2, 17), "排名")
        # self.ax = self.ax.resize((2000, int(self.ax.size[1] / self.ax.size[0] * 2000)), Image.ANTIALIAS)
        # self.ax = self.ax.quantize(colors=128, method=2)
        # 开始画数据
        for station_index, station_name, in enumerate(STATIONS_CNEMC.keys()):
            if station_name == "彩虹桥":
                self.draw_rec_text((station_index + 3, 0),
                                   station_index + 1,
                                   stroke_width=2, rfill="#ffff00")
                self.draw_rec_text((station_index + 3, 1),
                                   station_name,
                                   stroke_width=2, rfill="#ffff00")
            else:
                self.draw_rec_text((station_index + 3, 0),
                                   station_index + 1,)
                self.draw_rec_text((station_index + 3, 1),
                                   station_name)

        self.draw_rec_text((16, 0, 16, 0), "")
        self.draw_rec_text((16, 1, 16, 1), "全市")

        for species_index, species in enumerate(["PM25_RT", "PM25_RT_RANK", "PM25_DAY",  "PM25_DAY_RANK",
                                                "PM10_RT", "PM10_RT_RANK", "PM10_DAY", "PM10_DAY_RANK",
                                                 "NO2_RT", "NO2_RT_RANK", "NO2_DAY", "NO2_DAY_RANK",
                                                 "O3_RT", "O3_RT_RANK", "MDA8",  "MDA8_RANK"
                                                 ]):
            # 画国控点
            df_species = self.df.loc[list(
                STATIONS_CNEMC.keys())+["南京"], species].to_frame(name="VALUE")
            df_species["FONT"] = None
            df_species["RFILL"] = None

            df_species.loc["彩虹桥", "FONT"] = True
            df_species.loc["彩虹桥", "RFILL"] = "#ffff00"

            df_species = df_species.reset_index()

            for index, row_data in df_species.iterrows():
                if row_data.FONT:
                    self.draw_rec_text((index + 3, species_index + 2),
                                       row_data.VALUE,
                                       stroke_width=2, rfill="#ffff00"
                                       )
                else:
                    self.draw_rec_text(
                        (index + 3, species_index + 2),
                        row_data.VALUE)

            # 画城市
            # self.draw_rec_text((index + 16, species_index + 2), row_data.VALUE, fill=row_data.COLOR)
            # df_species = self.df.loc[["南京", "无锡", "苏州"],
            #                          species].to_frame(name="VALUE")
            # df_species["COLOR"] = "black"
            # df_species["FONT"] = None
            # df_species.loc["南京", "FONT"] = True
            # df_species["RFILL"] = None
            # df_species.loc["南京", "RFILL"] = "#ffff00"

            # df_species.loc[df_species["VALUE"] <
            #                df_species.at["南京", "VALUE"], "COLOR"] = "green"
            # df_species = df_species.reset_index()
            # for index, row_data in df_species.iterrows():
            #     print(index, species_index)
            #     if self.is_jn and row_data.FONT:
            #         self.draw_rec_text((index + 16, species_index + 2),
            #                            row_data.VALUE,
            #                            fill=row_data.COLOR,
            #                            font=True,
            #                            rfill=row_data.RFILL
            #                            )
            #     else:
            #         self.draw_rec_text(
            #             (index + 16, species_index + 2), row_data.VALUE, fill=row_data.COLOR)

        logger.info("绘制完整正在保存=>{}", output)
        self.ax = self.ax.resize(
            (2000, int(self.ax.size[1] / self.ax.size[0] * 2000)))
        self.ax.save(output)
        return output


class JiangningTextGenerator(GeneratorBase):
    def __init__(self, df,):
        super(JiangningTextGenerator, self).__init__(df)
        self.df = self.preprocessing(df)

    def preprocessing(self, df):
        df = df.set_index("DATETIME")
        df = df.rename({"PM2_5": "PM25"}, axis=1)
        realtime = df.loc[self.time_h.format("YYYY-MM-DD HH:00")]
        day = df.groupby("NAME").mean()
        day["MDA8"] = df.groupby("NAME")["O3_8H"].max()
        day.at["南京", "MDA8"] = day.loc[STATIONS_CNEMC.keys(), "MDA8"].mean()
        res_df = realtime.merge(day, left_on="NAME",
                                right_index=True, suffixes=["_RT", "_DAY"])
        res_df = res_df.round(0).set_index("NAME")
        rank = res_df.loc[STATIONS_CNEMC.keys()].rank(method="min").rename(
            lambda x: f"{x}_RANK", axis=1)
        res_df = res_df.join(rank)
        from aqi import AQI
        res_df[["AQI","AQI_RANK","EXCESS"]] = res_df.apply(lambda x: AQI().conc2aqi24h(x["PM25_DAY"], x["PM10_DAY"], np.nan, x["NO2_DAY"], x["MDA8"], np.nan),
                     axis=1,
                     result_type="expand"
                     )[["aqi", "rank","excess"]]

        return res_df

    @staticmethod
    def get_str(x, **kw):
        if isinstance(x, str):
            return x
        try:
            return str(round(x, **kw))
        except Exception as e:
            return "-"

    def run(self):
        if self.rt_count < 10:
            return False
        print(self.df.T)
        output = self.output_dir.joinpath(
            f"JN_{self.time_h.format('YYYY-MM-DDTHH')}.txt")

        from aqi import AQI        
        if self.time_h.hour in range(1, 17):
            res = f"南京市各国控站点{self.time_h.hour}时空气质量指标相关情况"

        elif self.time_h.hour in (0,17,18,19,20,21,22,23):
        # if 1:
            pm25_rt = self.get_str(self.df.at["彩虹桥", "PM25_RT"])
            pm25_rt_rank = self.get_str(self.df.at["彩虹桥", "PM25_RT_RANK"])
            pm25_day = self.get_str(self.df.at["彩虹桥", "PM25_DAY"])
            pm25_day_rank = self.get_str(self.df.at["彩虹桥", "PM25_DAY_RANK"])

            pm10_rt = self.get_str(self.df.at["彩虹桥", "PM10_RT"])
            pm10_rt_rank = self.get_str(self.df.at["彩虹桥", "PM10_RT_RANK"])
            pm10_day = self.get_str(self.df.at["彩虹桥", "PM10_DAY"])
            pm10_day_rank = self.get_str(self.df.at["彩虹桥", "PM10_DAY_RANK"])

            o3_rt = self.get_str(self.df.at["彩虹桥", "O3_RT"])
            o3_rt_rank = self.get_str(self.df.at["彩虹桥", "O3_RT_RANK"])
            o3_day = self.get_str(self.df.at["彩虹桥", "MDA8"])
            o3_day_rank = self.get_str(self.df.at["彩虹桥", "MDA8_RANK"])

            rank_text = (
                f"南京市各国控站点{self.time_h.hour}时空气质量指标相关情况，截止目前为止，"
                f"彩虹桥站点PM2.5实时浓度为{pm25_rt}微克/立方米，排名第{pm25_rt_rank}，"
                f"当日累计{pm25_day}微克/立方米，排名第{pm25_day_rank}； "
                f"PM10实时浓度为{pm10_rt}微克/立方米，排名第{pm10_rt_rank}，"
                f"当日累计{pm10_day}微克/立方米，排名第{pm10_day_rank}； "
                f"O3实时浓度为{o3_rt}微克/立方米，排名第{o3_rt_rank}，"
                f"当日累计{o3_day}微克/立方米，排名第{o3_day_rank}。"
            )
            print(self.df)

            # 17时后公共部分
            site_ex_jn = ['玄武湖', '瑞金路', '奥体中心', '草场门', '山西路', '迈皋桥', '仙林大学城', '中华门','浦口', '雄州', '永阳', '老职中']
            df_ex_jn = self.df.loc[site_ex_jn]
            site_all =  STATIONS_CNEMC.keys()
            df_site_all = self.df.loc[site_all]

            if self.time_h.hour != 23:
                df_site_all_polluted = df_site_all.query("AQI > 100")
                # 如果都为优良
                if df_site_all_polluted.empty:
                    # 有超标风险的站点
                    df_site_potential_polluted = df_site_all.query("70<PM25_DAY<75 or 150<MDA8<160")
                    #都没有风险的话
                    if df_site_potential_polluted.empty:
                        pst = f"彩虹桥国控站点空气质量等级为{self.df.at['彩虹桥','AQI_RANK']}，南京市空气质量等级为{self.df.at['南京','AQI_RANK']}。"
                    else:#有超标风险
                        # 彩虹桥有超标风险
                        pst = ""
                        for species in ("PM2.5","PM10","NO2","O3"):
                            if species == "PM2.5":
                                psite =  df_site_all.query("70<PM25_DAY<75")
                            elif species == "O3":
                                psite =  df_site_all.query("150<MDA8<160")
                            
                            if psite.empty:
                                ps = ""
                            else:
                                if "彩虹桥" in psite.index:
                                    text_psite = "、".join(psite.loc[psite.index!="彩虹桥"].index)[::-1].replace("、", "和", 1)[::-1]
                                    ps = f"彩虹桥站、{text_psite}站{species}有超标风险。"
                                else:# 彩虹桥没风险
                                    text_psite = "、".join(psite.index)[::-1].replace("、", "和", 1)[::-1]
                                    ps = f"{text_psite}站{species}有超标风险,彩虹桥站暂无风险。"
                            pst =  pst + ps
                else:#有站点超标
                    #彩虹桥超标？
  
                    #筛选出当前物种超标的站点
                    pst = ""
                    for species in ("PM2.5","PM10","NO2","O3"):
                        if species == "PM2.5":
                            psite =  df_site_all.query("PM25_DAY>75")
                        elif species == "PM10":
                            psite =  df_site_all.query("PM10_DAY>150")
                        elif species == "NO2":
                            psite =  df_site_all.query("NO2_DAY>80")
                        elif species == "O3":
                            psite =  df_site_all.query("MDA8>160")

                        if psite.empty:
                            ps = ""
                        else:
                            if "彩虹桥" in psite.index:
                                text_psite = "、".join(psite.loc[psite.index!="彩虹桥"].index)[::-1].replace("、", "和", 1)[::-1]
                                ps = f"彩虹桥站、{text_psite}站{species}已超标。"
                            else:# 其他站点超标，彩虹桥暂时没超标
                                text_psite = "、".join(psite.index)[::-1].replace("、", "和", 1)[::-1]
                                ps = f"{text_psite}站{species}已超标，彩虹桥站暂未超标。"

                                # 超标污染物为PM25,计算余量
                                if species == "PM2.5":
                                    jn_pm25 = self.df_ts.query("NAME=='彩虹桥'")["PM25"]
                                    try:
                                        import math
                                        x = math.ceil(75*(24-jn_pm25.isna().sum()) - jn_pm25.sum())/(24-self.time_h.hour)
                                    except Exception as e:
                                        x = "-"
                                    pm25_rest = f"彩虹桥PM2.5未来保良浓度为{x}微克/立方米。"
                                    ps =  f"{ps},{pm25_rest}"

                        pst = pst + ps

            if self.time_h.hour == 23:
                df_site_exjn_polluted = df_ex_jn.query("AQI > 100")
                # 如果不污染
                if df_site_exjn_polluted.empty:
                    df_aqi_rank_1 = df_site_all.query("AQI < 51")
                    df_aqi_rank_2 = df_site_all.query("50 < AQI < 101")
                    if df_aqi_rank_1.shape[0] == 0:
                        ps = "各国控站均为良。" 
                    elif df_aqi_rank_1.shape[0] < 7:
                        text_rank_1 = "、".join(df_aqi_rank_1.index)[::-1].replace("、", "和", 1)[::-1]
                        ps = f"{text_rank_1}为优，其他站点均为良。"
                    elif df_aqi_rank_1.shape[0] < 13:
                        text_rank_2 = "、".join(df_aqi_rank_2.index)[::-1].replace("、", "和", 1)[::-1]
                        ps = f"{text_rank_2}为良，其他站点均为优。"
                    else:
                        ps = "各国控站均为优。"
                else:
                    pst = ""
                    for species in ("PM2.5","PM10","NO2","O3"):
                        if species == "PM2.5":
                            psite =  df_site_all.query("PM25_DAY>75")
                        elif species == "PM10":
                            psite =  df_site_all.query("PM10_DAY>150")
                        elif species == "NO2":
                            psite =  df_site_all.query("NO2_DAY>80")
                        elif species == "O3":
                            psite =  df_site_all.query("MDA8>160")
                        
                        if psite.empty:
                            ps = ""
                        else:
                            if "彩虹桥" in psite.index:
                                text_psite = "、".join(psite.loc[psite.index!="彩虹桥"].index)[::-1].replace("、", "和", 1)[::-1]
                                ps = f"彩虹桥站、{text_psite}站{species}已超标。"
                            else:# 其他站点超标，彩虹桥暂时没超标
                                text_psite = "、".join(psite.index)[::-1].replace("、", "和", 1)[::-1]
                                ps = f"{text_psite}站{species}已超标，彩虹桥保良成功。" 
                        pst = pst + ps

            res =  rank_text + pst
        logger.info(res)
        output.write_text(res)
        return res


if __name__ == "__main__":
    from crawler import Cnemc,Moji

    c = Moji()
    c = Cnemc()
    df=c.run()
    # df.to_csv("test.csv")
    # df = pd.read_csv("test.csv", index_col=0)
    # df["DATETIME"] = pd.to_datetime(df.DATETIME)
    g = JiangningTextGenerator(df)
    print(g.rt_count)
    g.run()
