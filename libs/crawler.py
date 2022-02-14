import time
import numpy as np
import openpyxl
import arrow
import pandas as pd
import requests
from requests.adapters import HTTPAdapter
import json
from lxml import etree
from pathlib import Path
from loguru import logger
import io

from .vc import VC
from .config import STATIONS
from .cnemc import CNEMC


class Crawler:
    def __init__(self) -> None:
        self.rt = Path(__file__).parent.parent.absolute() / "rt"
        self.cookies_path = self.rt / Path(".cookies.json")
        self.validate_code_path = self.rt / Path("vc.png")
        self.rt_html_path = self.rt / Path("rt_data.html")
        self.datetime_tag_path = self.rt / Path("datetime.tag")
        self.rt_data_path = self.rt / Path("rt_data.csv")
        self.daily_data_path = self.rt / Path("daily_data.csv")
        self.all_data_path = self.rt / Path("all_data.csv")
        self.wuxi_data_path = self.rt / Path("wuxi_data.csv")
        self.suzhou_data_path = self.rt / Path("suzhou_data.csv")
        self.init_session()
        self.load_session()

    def init_session(self):
        self.session = requests.Session()
        self.session.headers = {
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Referer': 'http://112.25.188.53:12080/njeqs/mainFrame.aspx',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        }
        self.session.keep_alive = False
        self.session.mount('http://', HTTPAdapter(max_retries=3))
        self.session.mount('https://', HTTPAdapter(max_retries=3))

    def get_login_params(self):
        url = 'http://112.25.188.53:12080/njeqs/Default.aspx'
        logger.info("开始获取index页面")
        response = self.session.get(url, verify=False)
        html = etree.HTML(response.text)

        __VIEWSTATE = html.xpath("*//input[@id='__VIEWSTATE']/@value")
        __VIEWSTATEGENERATOR = html.xpath(
            "*//input[@id='__VIEWSTATEGENERATOR']/@value")
        __EVENTVALIDATION = html.xpath(
            "*//input[@id='__EVENTVALIDATION']/@value")

        if __VIEWSTATE and __VIEWSTATEGENERATOR and __EVENTVALIDATION:
            params = {
                "__VIEWSTATE": __VIEWSTATE[0],
                "__VIEWSTATEGENERATOR": __VIEWSTATEGENERATOR[0],
                "__EVENTVALIDATION": __EVENTVALIDATION[0]
            }
            logger.info("获取登录参数=>{}", params)
        else:
            logger.error("获取登录参数失败=>{}", response.text)

        data = {
            'userId': 'Daqs',
            'pwd': 'Daqs@123',
            'tbValidateCode': self.get_validate_code_word(),
            'loginBtn.x': '48',
            'loginBtn.y': '17',
            **params
        }
        return data

    def get_validate_code_word(self):
        # 获取验证码
        logger.info("正在获取验证码图片")
        url = "http://112.25.188.53:12080/njeqs/ValidateCodeCreate.aspx"
        img = self.session.get(url)
        if img.status_code == 200:
            logger.info("获取验证码图片成功,正在保存=>{}", self.validate_code_path)
        self.validate_code_path.write_bytes(img.content)
        logger.info("保存成功=>{}", self.validate_code_path)
        vc = VC()
        words = vc.get_words()
        return words

    def dump_session(self):
        logger.info("dumping session=>{}", self.cookies_path)
        with open(self.cookies_path, "w") as f:
            cookies_dict = requests.utils.dict_from_cookiejar(
                self.session.cookies)
            logger.info("获取到cookies为=>{}", cookies_dict)
            json.dump(cookies_dict, f)
        logger.info("保存会话成功=>{}", self.cookies_path)

    def load_session(self):
        if self.cookies_path.is_file():
            logger.info("发现cookies文件,正在恢复")
            with open(self.cookies_path, "r") as f:
                try:
                    cookies_dict = json.load(f)
                    cookies = requests.utils.cookiejar_from_dict(cookies_dict)
                    self.session.cookies = cookies
                    logger.info("恢复会话成功")
                except Exception as e:
                    logger.error("恢复会话失败=>{}", e)
                    logger.info("重新登录...")
                    self.login()
        else:
            logger.info("没有cookies,重新登录...")
            self.login()

    def login(self):
        for t in range(10):
            self.init_session()
            logger.info("第{}次登录..", t + 1)
            data = self.get_login_params()
            logger.info("登录数据为=>{}", data)
            logger.info("开始登录")
            login_url = 'http://112.25.188.53:12080/njeqs/Default.aspx'
            login_res = self.session.post(login_url, data=data)

            if login_res.status_code == 200 and "注消当前登录用户" in login_res.text:
                logger.info("登录成功")
                self.dump_session()
                break
            else:
                html = etree.HTML(login_res.text)
                error_msg = html.xpath("*//span[@id='hintTD']/text()")
                logger.error("登录失败=>{}", login_res.text)
                logger.error("错误信息为=>{}", error_msg)
                logger.error("等待十秒再试...")
                time.sleep(10)
        else:
            logger.info("十次登录失败,请及时处理...")
            from push import EmailPush
            push = EmailPush()
            push.mail("【空气质量速报】告警！！！", content="十次登录失败,请及时处理...")

    def get_rt_data(self):
        url = 'http://112.25.188.53:12080/njeqs/DataQuery/AirStationLastHourDataStat.aspx'
        params = {
            'strStationIds': 'Station.id in(2001,2002,2003,2004,2006,2007,2008,2009,2010,2016,2022,2024,2036)',
            'ftid': '578'
        }
        logger.info("开始获取数据,请求url为=>{}", url)
        response = self.session.post(url, params=params, timeout=60)
        logger.info("请求结果长度为=>{}", len(response.text))
        # 判断请求数据是否正常
        if "loginBtn" in response.text:
            logger.error("cookies失效,重新登录...")
            self.login()
            return self.get_rt_data()
        else:
            logger.info("获取数据成功可用")
            self.rt_html_path.write_text(response.text)
            logger.info("请求结果写入文件成功=>{}", self.rt_html_path)

    def get_datetime(self):
        html = etree.HTML(self.rt_html_path.read_text())
        datetime = html.xpath("*//input[@id='strStartTime']/@value")
        datetime = arrow.get(datetime[0]).shift(hours=1)
        return datetime

    def is_update(self):
        self.get_rt_data()
        new = self.get_datetime().format("YYYY-MM-DDTHH")

        if self.datetime_tag_path.is_file():
            last = self.datetime_tag_path.read_text()
        else:
            last = None

        return new != last

    def save_rt_data(self):
        df = pd.read_html(self.rt_html_path, encoding="utf-8",
                          attrs={'id': 'containerTB'})[0]

        df = df[["序号", "站点名称", "PM2.5(mg/m3)", "PM10(mg/m3)", "NO2(mg/m3)"]]
        df = df.rename({"序号": "ID", "站点名称": "STATION_NAME", "PM2.5(mg/m3)": "PM25",
                        "PM10(mg/m3)": "PM10", "NO2(mg/m3)": "NO2"}, axis=1)

        df.set_index("ID", inplace=True)
        df["PM25"] = df["PM25"] * 1000
        df["PM10"] = df["PM10"] * 1000
        df["NO2"] = df["NO2"] * 1000
        # 无效值剔除
        df = df.set_index("STATION_NAME")
        df = df.applymap(lambda x: float(x) if 0 < float(x) < 1000 else np.nan)
        df.loc[df["PM25"] > df["PM10"], "PM10"] = np.nan

        df = df.rename({"六合雄州": "雄州", "溧水永阳": "永阳",
                        "高淳老职中": "老职中", "江宁彩虹桥": "彩虹桥"})
        sites = ['玄武湖', '瑞金路', '奥体中心', '草场门', '山西路', '迈皋桥', '仙林大学城',
                 '中华门', '彩虹桥', '浦口', '雄州', '永阳', '老职中']
        df = df.loc[sites, :]
        df.to_csv(self.rt_data_path, float_format="%.0f")
        logger.info("rt_data写入成功=>{}", self.rt_data_path)
        return True

    def save_wuxi_data(self):
        service = CNEMC()
        try:
            res = service.get_city_history(city_code="320200")

            d = res["GetCityAQIPublishHistoriesResponse"]["GetCityAQIPublishHistoriesResult"]["RootResults"][
                "CityAQIPublishHistory"]
            df = pd.DataFrame(d)
            df = df.loc[:, ["TimePoint", "PM2_5", "PM10", "NO2", "O3", "CO", "SO2", "AQI"]]
            df = df.set_index("TimePoint")
            df.index = pd.to_datetime(df.index, format="%Y-%m-%dT%H:%M:%S")
            logger.info(df)
        except Exception as e:
            logger.error(e)
            return None

        data_hour = df.index.max().strftime("%Y%m%d%H")
        now_hour = arrow.now().shift(minutes=-27).format("YYYYMMDDHH")
        logger.info("data_hour=>{}", data_hour)
        logger.info("now_hour=>{}", now_hour)
        if data_hour == now_hour:
            df.to_csv(self.wuxi_data_path)
            return d

    def save_suzhou_data(self):
        service = CNEMC()
        try:
            res = service.get_city_history(city_code="320500")

            d = res["GetCityAQIPublishHistoriesResponse"]["GetCityAQIPublishHistoriesResult"]["RootResults"][
                "CityAQIPublishHistory"]
            df = pd.DataFrame(d)
            df = df.loc[:, ["TimePoint", "PM2_5", "PM10", "NO2", "O3", "CO", "SO2", "AQI"]]
            df = df.set_index("TimePoint")
            df.index = pd.to_datetime(df.index, format="%Y-%m-%dT%H:%M:%S")
            logger.info(df)
        except Exception as e:
            logger.error(e)
            return None

        data_hour = df.index.max().strftime("%Y%m%d%H")
        now_hour = arrow.now().shift(minutes=-27).format("YYYYMMDDHH")
        logger.info("data_hour=>{}", data_hour)
        logger.info("now_hour=>{}", now_hour)
        if data_hour == now_hour:
            df.to_csv(self.suzhou_data_path)
            return d

    def save_daily_data(self):
        dfs = []
        for station_name, station_id in STATIONS.items():
            logger.info("获取站点当日数据=>{},{}", station_name, station_id)
            station_df = self.get_station_data(station_id, station_name)
            dfs.append(station_df)
        logger.info("正在合并站点数据")
        all_station = pd.concat(dfs)
        daily_mean = all_station.groupby("STATION_NAME").mean()
        daily_mean.to_csv(self.daily_data_path, float_format="%.0f")
        logger.info("daily_data写入成功=>{}", self.daily_data_path)

    def save_daily_data2(self):
        # 首先获取请求参数
        url = 'http://112.25.188.53:12080/njeqs/DataQuery/AirStationDataStat.aspx'
        params = (
            ('strDimMonOptTypeID', '6'),
            ('strTimeGranularity', '20'),
            ('ftid', '603'),
        )
        data = {
            'ddlStationID': '2001',
        }
        response = self.session.post(url, params=params, data=data)
        html = etree.HTML(response.text)

        __VIEWSTATE = html.xpath("*//input[@id='__VIEWSTATE']/@value")
        __VIEWSTATEGENERATOR = html.xpath(
            "*//input[@id='__VIEWSTATEGENERATOR']/@value")
        __EVENTVALIDATION = html.xpath(
            "*//input[@id='__EVENTVALIDATION']/@value")

        if __VIEWSTATE and __VIEWSTATEGENERATOR and __EVENTVALIDATION:
            data_params = {
                "__VIEWSTATE": __VIEWSTATE[0],
                "__VIEWSTATEGENERATOR": __VIEWSTATEGENERATOR[0],
                "__EVENTVALIDATION": __EVENTVALIDATION[0]
            }
            logger.info("获取站点请求参数成功=>{}", data_params)
        else:
            logger.error("获取站点请求参数失败=>{}", response.text)
        # 开始遍历站点
        dfs = []
        for station_name, station_id in STATIONS.items():
            logger.info("获取站点当日数据=>{},{}", station_name, station_id)
            start = arrow.now().shift(hours=-1).floor("day").format("YYYY-MM-DD HH:00")
            end = arrow.now().shift(hours=-1).floor("day").shift(days=1).format("YYYY-MM-DD HH:00")
            logger.info(start, end)
            data = {
                **data_params,
                'ddlStationID': station_id,
                'strAuditData': '',
                'strAuditButton': '',
                'ddlDataType': '2',
                'strStartTime': start,
                'strEndTime': end,
                'btnQuery': '\u67E5 \u8BE2'
            }
            try:
                response = self.session.post('http://112.25.188.53:12080/njeqs/DataQuery/AirStationDataStat.aspx',
                                             params=params, data=data, timeout=10)
            except requests.exceptions.RequestException as e:
                logger.error("重试3次后仍未成功=>{}", e)
                return False

            html = etree.HTML(response.text)

            df = pd.read_html(io.StringIO(response.text),
                              encoding="utf-8", attrs={'id': 'tblContainer'})[0]

            datetime = self.get_datetime()
            history_file = self.rt.parent.joinpath("api").joinpath(
                f"{datetime.format('YYYY-MM-DDTHH')}_{station_name}.csv")
            df.to_csv(history_file,index=None)
            df = df.loc[:, ["时间", "PM2.5(mg/m3)", "PM10(mg/m3)", "NO2(mg/m3)"]]
            df = df.rename({"时间": "DATETIME",
                            "PM2.5(mg/m3)": "PM25_CUM",
                            "PM10(mg/m3)": "PM10_CUM",
                            "NO2(mg/m3)": "NO2_CUM"
                            }, axis=1)
            df = df.set_index("DATETIME")
            df.index = pd.to_datetime(df.index, format="%Y-%m-%d %H:%M", errors="coerce")
            df = df.loc[df.index.notna()]
            df = df.shift(periods=1, freq="H")
            # 无效值剔除
            df = df.applymap(lambda x: float(x) * 1000 if 0 < float(x) < 1.0 else np.nan)
            # PM25>PM10
            df.loc[df["PM25_CUM"] > df["PM10_CUM"], "PM10_CUM"] = np.nan
            df["STATION_NAME"] = station_name
            logger.info(df.index)
            dfs.append(df)

        logger.info("正在合并站点数据")
        all_station = pd.concat(dfs)
        daily_mean = all_station.groupby("STATION_NAME").mean()
        daily_mean.to_csv(self.daily_data_path, float_format="%.0f")
        logger.info("daily_data写入成功=>{}", self.daily_data_path)

    def get_station_data(self, station_id, station_name):
        url = "http://112.25.188.53:12080/njeqs/RTDataShow/AirHISDataShow_RTDB.aspx"
        params = {
            'strStationID': str(station_id)
        }
        logger.info("请求站点数据...")
        station_res = self.session.get(url, params=params, timeout=30)
        # (self.rt / Path(f"station_{station_id}.html")).write_text(station_res.text)
        datetime = self.get_datetime()
        df = pd.read_html(io.StringIO(station_res.text),
                          encoding="utf-8", attrs={'id': 'tblContainer'})[0]

        df = df.loc[:, ["时间", "PM2.5(mg/m3)", "PM10(mg/m3)", "NO2(mg/m3)"]].iloc[:-3]
        df = df.rename({"时间": "DATETIME",
                        "PM2.5(mg/m3)": "PM25_CUM",
                        "PM10(mg/m3)": "PM10_CUM",
                        "NO2(mg/m3)": "NO2_CUM"}, axis=1)
        df = df.set_index("DATETIME")
        df.index = pd.to_datetime(df.index, format="%Y-%m-%d %H:%M", errors="ignore")
        df = df.loc[df.index.minute == 0, :]
        df = df * 1000
        df["STATION_NAME"] = station_name
        logger.info("成功获取...")
        return df
    def save_txt(self,rt_df,daily_df,wu_data,suzhou_data):

        rt_df.loc["全市", :] = rt_df.mean()
        rt_df.loc["无锡", :] = wu_data.iloc[-1, :].values
        rt_df.loc["苏州", :] = suzhou_data.iloc[-1, :].values

        daily_df.loc["全市", :] = daily_df.mean()
        daily_df.loc["无锡", :] = wu_data.mean().values
        daily_df.loc["苏州", :] = suzhou_data.mean().values

        rts = rt_df.stack()
        rts.name = "实时"

        daily = daily_df.rename({"PM25_CUM": "PM25", "PM10_CUM": "PM10", "NO2_CUM": "NO2"}, axis=1)
        print(daily_df)
        dailys = daily.stack(dropna=False)
        dailys.name = "当日累计"
        print(rts,dailys)
        res = pd.concat([rts, dailys], axis=1,)
        resstr = res.loc[rt_df.index].applymap(lambda x: "" if np.isnan(x) else round(x))
        resstr.index.names = ["位置", "物种"]
        datetime = self.get_datetime()
        resstr.to_csv(f"history/{datetime.format('YYYY-MM-DDTHH')}.txt")

    def write_excel(self):
        # 先合并数据
        rt_df = pd.read_csv(self.rt_data_path, index_col=0)
        daily_df = pd.read_csv(self.daily_data_path, index_col=0)
        all_df = pd.concat([rt_df, daily_df], axis=1)
        all_df = all_df[["PM25", "PM25_CUM", "PM10", "PM10_CUM", "NO2", "NO2_CUM"]]
        # all_df.iloc[-2:,:] = np.nan
        datetime = self.get_datetime()
        wb = openpyxl.load_workbook(f"static/template3.xlsx")
        ws = wb["DATA"]

        ws["C1"].value = self.get_datetime().format("YYYY-MM-DD HH:mm")

        strdf = all_df.applymap(lambda x: int(x) if not np.isnan(x) else "-")
        for i, row in enumerate(ws["C4:H16"]):
            for j, col in enumerate(row):
                col.value = strdf.iloc[i, j]

        #无锡
        wu_data = pd.read_csv(self.wuxi_data_path, index_col=0, parse_dates=True, na_values=["-", "—", ""])
        wu_data = wu_data.loc[:, ["PM2_5", "PM10", "NO2"]]
        wu_day = wu_data.mean()

        ws["C18"] = wu_data.iloc[-1, 0]
        ws["D18"] = round(wu_day[0])
        ws["E18"] = wu_data.iloc[-1, 1]
        ws["F18"] = round(wu_day[1])
        ws["G18"] = wu_data.iloc[-1, 2]
        ws["H18"] = round(wu_day[2])

        #苏州
        suzhou_data = pd.read_csv(self.suzhou_data_path, index_col=0, parse_dates=True, na_values=["-", "—", ""])
        suzhou_data = suzhou_data.loc[:, ["PM2_5", "PM10", "NO2"]]
        suzhou_day = suzhou_data.mean()

        ws["C19"] = suzhou_data.iloc[-1, 0]
        ws["D19"] = round(suzhou_day[0])
        ws["E19"] = suzhou_data.iloc[-1, 1]
        ws["F19"] = round(suzhou_day[1])
        ws["G19"] = suzhou_data.iloc[-1, 2]
        ws["H19"] = round(suzhou_day[2])

        for i, row in enumerate(ws["C4:H16"]):
            for j, col in enumerate(row):
                col.value = strdf.iloc[i, j]
        excel_path = f"excel/{datetime.format('YYYY-MM-DDTHH')}.xlsx"
        logger.info("保存数据至=>{}", excel_path)
        wb.save(excel_path)

        # 保存数据去画图
        all_df.to_csv(self.all_data_path, float_format="%.0f")

        #保存文本格式

        self.save_txt(rt_df,daily_df,wu_data,suzhou_data)

    def run(self, force=False):
        if self.is_update() or force:
            datetime = self.get_datetime()
            logger.info("发现新数据=>{}", datetime)

            self.save_rt_data()
            self.save_daily_data2()
            if self.save_wuxi_data() and self.save_suzhou_data():
                logger.info("数据已完整")
                self.write_excel()
                self.datetime_tag_path.write_text(datetime.format("YYYY-MM-DDTHH"))
                return True
        else:
            logger.info("还没有新数据")
            return False


if __name__ == "__main__":
    crawer = Crawler()
    crawer.run()
    # crawer.get_all_station()
    # logger.info(crawer.get_dataframe())
